# -*- coding: utf-8 -*-
"""Excel ejecutivo PDNA para remisión al equipo sectorial.

Portada, leyenda, cruces territoriales/tipológicos y análisis 2.º nivel,
con formato de presentación (encabezados, filtros, números es-VE).
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pdna_costs import marco_pdna_ligero, matriz_pdna_completa, proyectar_pdna
from pdna_export import construir_export_pdna_fisico, matriz_fisica_desglosada
from pdna_salidas import construir_matriz_2do_nivel
from process_habitable import (
    ESQUEMA_PDNA_PISO_A_PISO,
    ETIQUETAS,
    aplicar_tipologia_pdna,
    etiqueta_display,
)

TZ = ZoneInfo("America/Caracas")

FILL_HEADER = PatternFill("solid", fgColor="1E3A5F")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
FILL_TITLE = PatternFill("solid", fgColor="0F2744")
FONT_TITLE = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
FILL_KPI = PatternFill("solid", fgColor="E8EEF5")
FONT_BODY = Font(name="Calibri", size=11)
FONT_MUTED = Font(name="Calibri", size=10, italic=True, color="475569")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

# Semáforo: relleno suave en encabezados de columnas de conteo
FILL_VERDE = PatternFill("solid", fgColor="166534")
FILL_AMARILLO = PatternFill("solid", fgColor="A16207")
FILL_ROJO = PatternFill("solid", fgColor="991B1B")
FILL_NEGRO = PatternFill("solid", fgColor="111827")

USO_LABEL = {
    "casa": "Casa",
    "edificio": "Edificio",
    "turismo": "Establecimientos turísticos",
    "comercio": "Comercio",
    "oficina": "Comercio",
    "otros": "Otros",
}


def _fmt_es_int(n: float | int) -> str:
    return f"{int(round(float(n))):,}".replace(",", ".")


def _label_uso(s: Any) -> str:
    t = str(s or "").strip().lower()
    if t in USO_LABEL:
        return USO_LABEL[t]
    raw = str(s or "").strip()
    return raw or "—"


def _label_pisos(s: Any) -> str:
    t = str(s or "").strip()
    if not t or t.lower() in {"nan", "none", "(sin dato)", "pisos s/d", "s/d"}:
        return "Sin dato de pisos"
    return t


def _rename_semaforo_cols(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    ren = {
        "VERDE": "Verde",
        "AMARILLO": "Amarillo",
        "ROJO": "Rojo",
        "NEGRO": "Pérdida total",
        "Negro_perdida_total": "Pérdida total",
        "Negro": "Pérdida total",
        "verde": "Verde",
        "amarillo": "Amarillo",
        "rojo": "Rojo",
        "negro": "Pérdida total",
    }
    return out.rename(columns={k: v for k, v in ren.items() if k in out.columns})


def _humanize_export_sheets(sheets: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    for name, df in sheets.items():
        if df is None or df.empty:
            continue
        d = _rename_semaforo_cols(df)
        if "Uso" in d.columns:
            d["Uso"] = d["Uso"].map(_label_uso)
        if "Pisos" in d.columns:
            d["Pisos"] = d["Pisos"].map(_label_pisos)
        if "Banda_pisos" in d.columns:
            d = d.rename(columns={"Banda_pisos": "Pisos"})
            d["Pisos"] = d["Pisos"].map(_label_pisos)
        if "Tipo_dano" in d.columns:
            d = d.rename(columns={"Tipo_dano": "Tipo de daño"})
        if "Semaforo" in d.columns:
            d["Semaforo"] = d["Semaforo"].map(
                lambda x: etiqueta_display(x) if str(x).upper() in ETIQUETAS else x
            )
            d = d.rename(columns={"Semaforo": "Semáforo"})
        if "Codigo" in d.columns:
            d = d.drop(columns=["Codigo"], errors="ignore")
        out[name] = d
    return out


def _resumen_ejecutivo(work: pd.DataFrame, *, esquema: str) -> pd.DataFrame:
    tip = work["tipologia_pdna"].notna() & work["etiqueta_n"].isin(ETIQUETAS)
    sub = work.loc[tip]
    et = sub["etiqueta_n"].astype(str).str.upper()
    n = int(len(sub))
    n_rojo = int((et == "ROJO").sum())
    n_negro = int((et == "NEGRO").sum())
    n_crit = n_rojo + n_negro
    n_tur = int((work.get("uso_grupo") == "Establecimientos turísticos").sum()) if "uso_grupo" in work.columns else 0
    ahora = datetime.now(TZ).strftime("%d/%m/%Y %H:%M")
    rows = [
        ("Documento", "Insumo físico PDNA — habitabilidad post-evento", ""),
        ("Destinatario", "Equipo sectorial PDNA / recuperación", ""),
        ("Generado", ahora, "America/Caracas"),
        ("Universo de inspecciones", _fmt_es_int(len(work)), "Corte Habitable cargado en el BI"),
        ("Unidades con tipología PDNA", _fmt_es_int(n), "Material × uso × pisos + semáforo válido"),
        ("Verde", _fmt_es_int(int((et == "VERDE").sum())), "Habitable — acceso permitido"),
        ("Amarillo", _fmt_es_int(int((et == "AMARILLO").sum())), "Habitable — acceso restringido"),
        ("Rojo", _fmt_es_int(n_rojo), "No habitable / alto riesgo"),
        ("Pérdida total", _fmt_es_int(n_negro), "Etiqueta negra"),
        ("Daño crítico (Rojo + pérdida total)", _fmt_es_int(n_crit), f"{100.0 * n_crit / max(n, 1):.1f}% de tipificadas"),
        ("Establecimientos turísticos (uso)", _fmt_es_int(n_tur), "Detectados en uso, nombre, observaciones o dirección"),
        ("Esquema de tipología", "Piso a piso (1–20 + 21 o más)" if esquema == ESQUEMA_PDNA_PISO_A_PISO else esquema, "Valores >60 pisos = sin dato"),
        ("Moneda / costos", "Este archivo es físico (conteos)", "Sin estimación USD; los costos se calibran fuera"),
        (
            "Nota",
            "Las cifras son insumos de trabajo hasta calibración sectorial.",
            "No sustituyen el dictamen de campo ni la valoración oficial.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Indicador", "Valor", "Lectura"])


def _leyenda() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("0. Resumen", "Cifras clave del corte para la mesa PDNA."),
            ("1. Cómo leer", "Esta hoja: significado de pestañas y campos."),
            ("2. Por territorio", "Conteo por Estado / Municipio / Parroquia × material × uso × pisos × tipo de daño × semáforo."),
            ("3. Por tipología", "Misma desagregación sin territorio (material × uso × pisos × tipo de daño)."),
            ("4. Matriz tipología", "Vista agregada tipología constructiva × semáforo (una fila = una combinación observada)."),
            ("5. Totales semáforo", "Suma de unidades tipificadas por color."),
            ("6. Totales tipo de daño", "Estructural / no estructural / ambos / sin daño relevante."),
            ("7. Uso × semáforo", "Supercategoría de uso (incl. turismo y comercio) frente al semáforo."),
            ("8. Análisis 2.º nivel", "Territorio × tipología de inmueble × n.º de pisos × estratificación de daño."),
            ("9. Resumen estratificación", "Totales del análisis 2.º nivel por categoría de estratificación."),
            ("Campo · Uso", "Casa, Edificio, Establecimientos turísticos, Comercio (oficina incluida en Comercio)."),
            ("Campo · Pisos", "Piso a piso de 1 a 20; cola «21 o más»; sin dato si falta o es implausible (>60)."),
            ("Campo · Semáforo", "Verde / Amarillo / Rojo / Pérdida total (antes «Negro»)."),
            ("Campo · Tipo de daño", "Clasificación ANIH: estructural, no estructural, ambos o sin daño relevante."),
        ],
        columns=["Elemento", "Significado"],
    )


def _uso_x_semaforo(work: pd.DataFrame) -> pd.DataFrame:
    if "uso_grupo" not in work.columns or "etiqueta_n" not in work.columns:
        return pd.DataFrame()
    tip = work["tipologia_pdna"].notna() & work["etiqueta_n"].isin(ETIQUETAS)
    sub = work.loc[tip].copy()
    if sub.empty:
        sub = work.loc[work["etiqueta_n"].isin(ETIQUETAS)].copy()
    ct = (
        pd.crosstab(sub["uso_grupo"], sub["etiqueta_n"])
        .reindex(columns=list(ETIQUETAS), fill_value=0)
        .reset_index()
        .rename(columns={"uso_grupo": "Uso"})
    )
    ct = _rename_semaforo_cols(ct)
    if "Uso" in ct.columns:
        ct["Uso"] = ct["Uso"].map(_label_uso)
    ct["Total"] = ct[[c for c in ("Verde", "Amarillo", "Rojo", "Pérdida total") if c in ct.columns]].sum(axis=1)
    return ct.sort_values("Total", ascending=False).reset_index(drop=True)


def _matriz_2n_ejecutiva(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    m2 = construir_matriz_2do_nivel(df)
    if m2.empty:
        return m2, pd.DataFrame()
    show = m2.rename(
        columns={
            "estado": "Estado",
            "municipio": "Municipio",
            "parroquia": "Parroquia",
            "tipologia": "Tipología de inmueble",
            "num_pisos": "N.º pisos",
            "etiqueta": "Semáforo",
            "estratificacion": "Estratificación",
            "elementos_estructurales": "Elementos estructurales",
            "inspecciones": "Inspecciones",
            "personas": "Personas",
        }
    )
    if "Tipología de inmueble" in show.columns:
        show["Tipología de inmueble"] = show["Tipología de inmueble"].map(_label_uso)
    if "Semáforo" in show.columns:
        show["Semáforo"] = show["Semáforo"].map(lambda x: etiqueta_display(str(x).upper()))
    res = (
        show.groupby("Estratificación", dropna=False)
        .agg(Inspecciones=("Inspecciones", "sum"), Personas=("Personas", "sum"))
        .reset_index()
        .sort_values("Inspecciones", ascending=False)
    )
    return show, res


def construir_libro_pdna_ejecutivo(
    df: pd.DataFrame,
    *,
    esquema: str = ESQUEMA_PDNA_PISO_A_PISO,
) -> dict[str, pd.DataFrame]:
    """Arma el diccionario de hojas en orden ejecutivo."""
    work = aplicar_tipologia_pdna(df, esquema=esquema, copy=True)
    fis = _humanize_export_sheets(construir_export_pdna_fisico(work))

    lig = marco_pdna_ligero(work)
    lig = aplicar_tipologia_pdna(lig, esquema=esquema, copy=False)
    proj = proyectar_pdna(lig, slim=False)
    mat = matriz_pdna_completa(proj, solo_observadas=True, incluir_filas_vacias=False)
    vista = matriz_fisica_desglosada(mat)
    if not vista.empty:
        vista = _rename_semaforo_cols(vista)
        if "Uso" in vista.columns:
            vista["Uso"] = vista["Uso"].map(_label_uso)
        if "Pisos" in vista.columns:
            vista["Pisos"] = vista["Pisos"].map(_label_pisos)
        # Quitar fila TOTAL del cuerpo si existe (va al pie visual en Excel vía estilos)
        vista = vista.loc[vista["Material"].astype(str).str.upper() != "TOTAL"].copy()

    m2, est = _matriz_2n_ejecutiva(work)

    book: dict[str, pd.DataFrame] = {
        "0. Resumen": _resumen_ejecutivo(work, esquema=esquema),
        "1. Cómo leer": _leyenda(),
    }
    if "Por_territorio" in fis and not fis["Por_territorio"].empty:
        book["2. Por territorio"] = fis["Por_territorio"]
    if "Por_tipologia" in fis and not fis["Por_tipologia"].empty:
        book["3. Por tipología"] = fis["Por_tipologia"]
    if not vista.empty:
        book["4. Matriz tipología"] = vista
    if "Totales_semaforo" in fis and not fis["Totales_semaforo"].empty:
        book["5. Totales semáforo"] = fis["Totales_semaforo"]
    if "Totales_tipo_dano" in fis and not fis["Totales_tipo_dano"].empty:
        td = fis["Totales_tipo_dano"].rename(
            columns={"Tipo_dano": "Tipo de daño", "Unidades": "Unidades"}
        )
        book["6. Totales tipo de daño"] = td
    uso = _uso_x_semaforo(work)
    if not uso.empty:
        book["7. Uso × semáforo"] = uso
    if not m2.empty:
        book["8. Análisis 2.º nivel"] = m2
    if not est.empty:
        book["9. Resumen estratificación"] = est
    return book


def _style_workbook(path_or_buf: Path | BytesIO) -> None:
    from openpyxl import load_workbook

    if isinstance(path_or_buf, BytesIO):
        path_or_buf.seek(0)
    wb = load_workbook(path_or_buf)
    for ws in wb.worksheets:
        # Título en primera fila si es Resumen / Cómo leer
        is_meta = ws.title.startswith(("0.", "1."))
        header_row = 1
        for cell in ws[header_row]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = THIN
            val = str(cell.value or "")
            if val == "Verde":
                cell.fill = FILL_VERDE
            elif val == "Amarillo":
                cell.fill = FILL_AMARILLO
            elif val == "Rojo":
                cell.fill = FILL_ROJO
            elif val in {"Pérdida total", "Negro"}:
                cell.fill = FILL_NEGRO

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 22

        # Anchos
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = 12
            for cell in col[:80]:
                if cell.value is None:
                    continue
                width = max(width, min(42, len(str(cell.value)) + 2))
                if not is_meta and cell.row > 1:
                    cell.font = FONT_BODY
                    cell.border = THIN
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
            ws.column_dimensions[letter].width = width

        if is_meta:
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 40)):
                for cell in row:
                    cell.font = FONT_BODY
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                ws.row_dimensions[row[0].row].height = 18
            if ws.title.startswith("0."):
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.fill = FILL_KPI
                        cell.font = Font(name="Calibri", size=11, bold=True)

    if isinstance(path_or_buf, Path):
        wb.save(path_or_buf)
    else:
        path_or_buf.seek(0)
        wb.save(path_or_buf)
        path_or_buf.seek(0)


def escribir_excel_pdna_ejecutivo(
    df: pd.DataFrame,
    path: Path | str,
    *,
    esquema: str = ESQUEMA_PDNA_PISO_A_PISO,
) -> Path:
    """Escribe el Excel ejecutivo en disco y aplica formato."""
    path = Path(path)
    book = construir_libro_pdna_ejecutivo(df, esquema=esquema)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, frame in book.items():
            safe = str(name)[:31]
            frame.to_excel(xw, sheet_name=safe, index=False)
    _style_workbook(path)
    return path


def bytes_excel_pdna_ejecutivo(
    df: pd.DataFrame,
    *,
    esquema: str = ESQUEMA_PDNA_PISO_A_PISO,
) -> bytes:
    """Bytes del Excel ejecutivo (para botón Streamlit)."""
    book = construir_libro_pdna_ejecutivo(df, esquema=esquema)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, frame in book.items():
            frame.to_excel(xw, sheet_name=str(name)[:31], index=False)
    _style_workbook(buf)
    return buf.getvalue()
