# -*- coding: utf-8 -*-
"""Hojas Control_2da_ronda + Vaciado informe digital (listas desplegables)."""
from __future__ import annotations

from copy import copy
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


LISTAS = {
    "si_no_parcial": ["Sí", "No", "Parcial / corregir", "Pendiente"],
    "si_no_insuf": ["Sí", "No", "Insuficiente evidencia", "Pendiente"],
    "estado_2da": [
        "Pendiente verificación",
        "En visita",
        "Borrador",
        "Revisado",
        "Aprobado",
        "Publicado",
    ],
    "decision_D": [
        "D1 — Complementos requeridos",
        "D2 — Reparar / reconstruir",
        "D3 — Demoler",
        "D4 — Escombros / ya colapsado",
        "Pendiente",
    ],
    "magnitud_M": [
        "N/A (no es D2)",
        "M1 — Reparación local (menor intervención)",
        "M2 — Reparación importante",
        "M3 — Reconstrucción parcial",
        "M4 — Reconstrucción / refuerzo mayor (mayor intervención)",
        "Pendiente",
    ],
    "prioridad": ["Inmediata", "Alta", "Programable", "Pendiente"],
    "abc": ["A", "B", "C", "No observable", "Pendiente"],
    "ocupacion": [
        "Desalojado",
        "Ocupación parcial",
        "Habitado irregularmente",
        "Escombros",
        "Pendiente",
    ],
    "sistema": [
        "Pórticos concreto armado",
        "Muros de concreto",
        "Mixto",
        "Acero",
        "Mampostería",
        "Otro",
        "Pendiente",
    ],
    "pct_columnas": [
        "No observable",
        "<10%",
        "10–30%",
        ">30%",
        ">50%",
        "Pendiente",
    ],
    "inclinacion": [
        "No",
        "Sí — cualitativa",
        "Sí — con medición Δ",
        "No observable",
        "Pendiente",
    ],
    "peligro_aledanos": ["Sí", "No", "No observable", "Pendiente"],
}


def build_control_2da_ronda(
    ranking_lg: pd.DataFrame,
    df_cruce: pd.DataFrame,
) -> pd.DataFrame:
    """Cola de control: Alta/Muy alta La Guaira + casos del lote PDF."""
    ids_lote = set()
    if "hab_id" in df_cruce.columns:
        for v in df_cruce["hab_id"]:
            try:
                if pd.notna(v) and str(v).strip() not in ("", "nan"):
                    ids_lote.add(int(float(v)))
            except (TypeError, ValueError):
                pass

    base = ranking_lg.copy()
    mask = (
        base["banda_prioridad"].isin(["Muy alta", "Alta"])
        | base["id"].isin(ids_lote)
        | (base["score_gravedad"] >= 50)
    )
    ctrl = base.loc[mask].copy()
    # limitar a 400 para operatividad; priorizar score
    ctrl = ctrl.sort_values("score_gravedad", ascending=False).head(400)

    ctrl["estado_2da_ronda"] = "Pendiente verificación"
    ctrl["validacion_precarga"] = "Pendiente"
    ctrl["etiqueta_roja_confirmada"] = "Pendiente"
    ctrl["decision_D"] = "Pendiente"
    ctrl["magnitud_M"] = "Pendiente"
    ctrl["prioridad_operativa"] = "Pendiente"
    ctrl["fecha_visita_2"] = ""
    ctrl["evaluador_visita_2"] = ""
    ctrl["supervisor_visita_2"] = ""
    ctrl["resumen_ejecutivo"] = ""
    ctrl["correcciones_precarga"] = ""
    ctrl["n_fotos"] = ""

    # ejemplo Franco Mar ya dictaminado en el anexo del Word
    fm = ctrl["id"] == 171818
    if fm.any():
        ctrl.loc[fm, "estado_2da_ronda"] = "Revisado"
        ctrl.loc[fm, "validacion_precarga"] = "Parcial / corregir"
        ctrl.loc[fm, "etiqueta_roja_confirmada"] = "Sí"
        ctrl.loc[fm, "decision_D"] = "D3 — Demoler"
        ctrl.loc[fm, "magnitud_M"] = "N/A (no es D2)"
        ctrl.loc[fm, "prioridad_operativa"] = "Inmediata"
        ctrl.loc[fm, "fecha_visita_2"] = "2026-08-11"
        ctrl.loc[fm, "evaluador_visita_2"] = "Luis Burgos; José García"
        ctrl.loc[fm, "supervisor_visita_2"] = "Aura Quintero"
        ctrl.loc[fm, "correcciones_precarga"] = "Sótanos 1→2; precisar GPS/dirección de control"
        ctrl.loc[fm, "resumen_ejecutivo"] = (
            "Fase 1 ROJO score 51 (Media). Visita 11/08: piso crítico Nivel 1, "
            ">50% columnas graves, pérdida de verticalidad. Decisión D1 demoler."
        )

    keep = [
        "rank_gravedad_LaGuaira",
        "rank_gravedad",
        "score_gravedad",
        "banda_prioridad",
        "prob_relativa_demolicion",
        "en_lote_informes_pdf",
        "estado_2da_ronda",
        "validacion_precarga",
        "etiqueta_roja_confirmada",
        "decision_D",
        "magnitud_M",
        "prioridad_operativa",
        "fecha_visita_2",
        "evaluador_visita_2",
        "supervisor_visita_2",
        "id",
        "certificado",
        "nombre_edificacion",
        "direccion",
        "municipio",
        "parroquia",
        "num_pisos",
        "piso_critico",
        "riesgo_externo",
        "riesgo_severo",
        "ext_colapso_estructura",
        "acc_medidas",
        "score_detalle",
        "correcciones_precarga",
        "resumen_ejecutivo",
        "n_fotos",
        "lat",
        "lng",
        "inspector_nombre",
        "created_at",
    ]
    keep = [c for c in keep if c in ctrl.columns]
    return ctrl[keep].reset_index(drop=True)


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _write_df_sheet(wb, name: str, df: pd.DataFrame, freeze=True):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _style_header(ws)
    if freeze:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(28, max(12, len(str(col)) + 2))
    return ws


def _add_list_sheet(wb):
    name = "Listas_desplegables"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "listas_para_validacion"
    col = 1
    for key, values in LISTAS.items():
        ws.cell(1, col, key)
        for i, v in enumerate(values, start=2):
            ws.cell(i, col, v)
        col += 1
    ws.sheet_state = "hidden"
    return ws


def _dv(formula: str, sqref: str) -> DataValidation:
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valor no válido",
        error="Elija un valor de la lista.",
    )
    dv.add(sqref)
    return dv


def _section_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def write_vaciado_informe_sheet(
    wb,
    sheet_name: str,
    precarga: dict[str, Any],
    valores_visita: dict[str, Any] | None = None,
    titulo: str = "VACIADO DIGITAL — INSPECCIÓN DETALLADA (2.ª RONDA)",
):
    """Formulario vertical tipo informe: col A etiqueta, B valor (con listas)."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    valores_visita = valores_visita or {}

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    font_sec = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_lab = Font(name="Calibri", size=10, bold=True)
    font_val = Font(name="Calibri", size=10)
    font_note = Font(name="Calibri", size=9, italic=True, color="666666")

    ws["A1"] = titulo
    ws["A1"].font = font_title
    ws.merge_cells("A1:B1")
    ws["A2"] = (
        "Campos grises = precarga Fase 1 / ranking (validar). "
        "Campos blancos = captura de la visita detallada. Use listas desplegables donde existan."
    )
    ws["A2"].font = font_note
    ws.merge_cells("A2:B2")

    # (seccion, etiqueta, clave_valor, lista_key|None, es_precarga)
    rows_def = [
        ("A0 — PRECARGA FASE 1", None, None, None, False),
        ("A0", "ID Habitable", "hab_id", None, True),
        ("A0", "Certificado", "certificado", None, True),
        ("A0", "Nombre (Habitable)", "nombre_hab", None, True),
        ("A0", "Etiqueta Fase 1", "etiqueta_f1", None, True),
        ("A0", "Fecha inspección Fase 1", "fecha_f1", None, True),
        ("A0", "Inspector Fase 1", "inspector_f1", None, True),
        ("A0", "Dirección Habitable", "direccion_hab", None, True),
        ("A0", "Municipio / Parroquia", "muni_parr", None, True),
        ("A0", "Pisos / sótanos (Fase 1)", "pisos_f1", None, True),
        ("A0", "Riesgo externo / severo", "riesgos_f1", None, True),
        ("A0", "Colapso estructura (Fase 1)", "colapso_f1", None, True),
        ("A0", "Piso crítico (Fase 1)", "piso_crit_f1", None, True),
        ("A0", "Acciones Fase 1", "acciones_f1", None, True),
        ("A0", "Observaciones Fase 1", "obs_f1", None, True),
        ("A0", "GPS Habitable", "gps_hab", None, True),
        ("A0 — RANKING", None, None, None, False),
        ("A0", "Score gravedad (0–100)", "score", None, True),
        ("A0", "Banda prioridad", "banda", None, True),
        ("A0", "Puesto La Guaira / nacional", "puestos", None, True),
        ("A0", "Detalle del score", "score_detalle", None, True),
        ("A0", "Probabilidad relativa (texto)", "prob_rel", None, True),
        ("A0 — VALIDACIÓN", None, None, None, False),
        ("A0", "¿Edificio precargado correcto?", "val_edificio", "si_no_parcial", False),
        ("A0", "¿Etiqueta ROJA se mantiene?", "val_etiqueta", "si_no_insuf", False),
        ("A0", "¿Datos geométricos correctos?", "val_geometria", "si_no_parcial", False),
        ("A0", "¿Ranking útil como prioridad?", "val_ranking", "si_no_parcial", False),
        ("A0", "Correcciones a la precarga", "correcciones", None, False),
        ("A/B — IDENTIDAD Y EDIFICIO (VISITA 2)", None, None, None, False),
        ("A", "Nombre confirmado", "nombre_conf", None, False),
        ("A", "Fecha visita detallada", "fecha_v2", None, False),
        ("A", "GPS control (visita)", "gps_v2", None, False),
        ("A", "Evaluadores visita 2", "evaluadores_v2", None, False),
        ("A", "Supervisor", "supervisor_v2", None, False),
        ("B", "Uso", "uso", None, False),
        ("B", "Pisos confirmados", "pisos_conf", None, False),
        ("B", "Sótanos confirmados", "sotanos_conf", None, False),
        ("B", "Sistema estructural", "sistema", "sistema", False),
        ("B", "Ocupación actual", "ocupacion", "ocupacion", False),
        ("B", "Peligro aledaños / vía", "peligro_aledanos", "peligro_aledanos", False),
        ("C — DAÑO DETALLADO", None, None, None, False),
        ("C", "Piso crítico (visita 2)", "piso_crit_v2", None, False),
        ("C", "% columnas daño grave", "pct_columnas", "pct_columnas", False),
        ("C", "Inclinación / verticalidad", "inclinacion", "inclinacion", False),
        ("C", "Δ inclinación (si midió)", "delta_incl", None, False),
        ("C", "Daño vigas / losas", "dano_vigas", "abc", False),
        ("C", "Riesgo fachada (A/B/C)", "riesgo_fachada", "abc", False),
        ("C", "Análisis libre (ingeniero)", "analisis_libre", None, False),
        ("D — DECISIÓN DE CONTROL", None, None, None, False),
        ("D", "Estado del caso", "estado_2da", "estado_2da", False),
        ("D", "Decisión D1–D5", "decision_D", "decision_D", False),
        ("D", "Magnitud reparación M1–M4", "magnitud_M", "magnitud_M", False),
        ("D", "Prioridad operativa", "prioridad", "prioridad", False),
        ("D", "Medidas inmediatas", "medidas", None, False),
        ("D", "Justificación libre", "justificacion", None, False),
        ("E — EVIDENCIA", None, None, None, False),
        ("E", "N.º de fotos", "n_fotos", None, False),
        ("E", "Firmas (elaboró / revisó / aprobó)", "firmas", None, False),
        ("RESUMEN EJECUTIVO", None, None, None, False),
        ("R", "Resumen ejecutivo (≤1200 caracteres)", "resumen_ejecutivo", None, False),
    ]

    # merge precarga + visita defaults
    data = {**precarga, **valores_visita}

    r = 4
    dropdown_cells: list[tuple[int, str]] = []  # (row, lista_key)
    for seccion, etiqueta, key, lista_key, es_precarga in rows_def:
        if etiqueta is None:
            ws.cell(r, 1, seccion)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell = ws.cell(r, 1)
            cell.fill = _section_fill("2E75B6")
            cell.font = font_sec
            cell.alignment = Alignment(vertical="center")
            r += 1
            continue
        ws.cell(r, 1, etiqueta).font = font_lab
        ws.cell(r, 1).border = thin
        val = data.get(key, "")
        if val is None or (isinstance(val, float) and pd.isna(val)):
            val = ""
        c = ws.cell(r, 2, val)
        c.font = font_val
        c.border = thin
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if es_precarga:
            c.fill = PatternFill("solid", fgColor="D9E2F3")
        if lista_key:
            dropdown_cells.append((r, lista_key))
            c.fill = PatternFill("solid", fgColor="FFF2CC") if not es_precarga else c.fill
        # tall rows for free text
        if key in ("analisis_libre", "justificacion", "resumen_ejecutivo", "correcciones", "obs_f1"):
            ws.row_dimensions[r].height = 60
        r += 1

    # named ranges / validations referencing Listas_desplegables
    # Map lista keys to columns on Listas sheet
    list_cols = {k: i + 1 for i, k in enumerate(LISTAS.keys())}
    for row_i, lista_key in dropdown_cells:
        col_i = list_cols[lista_key]
        col_letter = get_column_letter(col_i)
        nvals = len(LISTAS[lista_key])
        formula = f"Listas_desplegables!${col_letter}$2:${col_letter}${nvals + 1}"
        ws.add_data_validation(_dv(formula, f"$B${row_i}"))

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 78
    ws.row_dimensions[1].height = 22
    return ws


def francomar_precarga() -> dict[str, Any]:
    return {
        "hab_id": "171818",
        "certificado": "73420200720261338",
        "nombre_hab": "Edif. Francomar",
        "etiqueta_f1": "ROJO",
        "fecha_f1": "2026-07-20",
        "inspector_f1": "Maria Mercedes Nieves Duarte",
        "direccion_hab": "Calle Principal Boulevard Tanaguarenas. Caraballeda. La Guaira.",
        "muni_parr": "Vargas / Caraballeda",
        "pisos_f1": "12 / 1",
        "riesgos_f1": "Externo C / Severo A",
        "colapso_f1": "C",
        "piso_crit_f1": "Planta baja, piso 1, 2, 3, 4",
        "acciones_f1": "Acordonar",
        "obs_f1": "Edificio a punto de colapso, se recomienda Desalojar",
        "gps_hab": "10.6123543, -66.8318071",
        "score": 51,
        "banda": "Media",
        "puestos": "La Guaira ~115 / Nacional ~264",
        "score_detalle": (
            "riesgo_externo+20; ext_colapso+15; acordonar+2; altura_10++6; "
            "piso_critico+3; sev_elementos+5"
        ),
        "prob_rel": "Media — verificar; puede ser ROJO por riesgo localizado",
    }


def francomar_visita() -> dict[str, Any]:
    return {
        "val_edificio": "Sí",
        "val_etiqueta": "Sí",
        "val_geometria": "Parcial / corregir",
        "val_ranking": "Sí",
        "correcciones": "Sótanos 1→2; precisar GPS y dirección de control de la visita",
        "nombre_conf": "Edificio Franco Mar",
        "fecha_v2": "2026-08-11",
        "gps_v2": "10.489497, -66.899001",
        "evaluadores_v2": "Ing. Luis Burgos; Ing. José García",
        "supervisor_v2": "Ing. Aura Quintero",
        "uso": "Vivienda edificio",
        "pisos_conf": "12",
        "sotanos_conf": "2",
        "sistema": "Pórticos concreto armado",
        "ocupacion": "Desalojado",
        "peligro_aledanos": "Sí",
        "piso_crit_v2": "Nivel 1 / PB transición",
        "pct_columnas": ">50%",
        "inclinacion": "Sí — cualitativa",
        "delta_incl": "",
        "dano_vigas": "C",
        "riesgo_fachada": "C",
        "analisis_libre": (
            "Piso crítico en Nivel 1 con >50% columnas graves; pérdida de verticalidad; "
            "daño en vigas en varios niveles. Ranking Media (51) no refleja la gravedad "
            "vista en sitio. Reparación en sitio no viable con seguridad."
        ),
        "estado_2da": "Revisado",
        "decision_D": "D3 — Demoler",
        "magnitud_M": "N/A (no es D2)",
        "prioridad": "Inmediata",
        "medidas": "Exclusión total; acordonar; no ingreso; monitoreo de vecinos",
        "justificacion": (
            "Se recomienda demolición controlada por falla masiva en piso crítico, "
            "pérdida de verticalidad y peligro a colindantes."
        ),
        "n_fotos": "Según informe detallado DGPS (registro fotográfico)",
        "firmas": "Elaboró: equipo visita 11/08 · Supervisión: Aura Quintero",
        "resumen_ejecutivo": (
            "Franco Mar (Caraballeda). Fase 1 ROJO 20/07/2026, score 51 (Media, ~puesto 115 "
            "La Guaira). Visita 11/08/2026: 12 pisos + 2 sótanos, piso crítico Nivel 1, "
            ">50% columnas graves, pérdida de verticalidad. Validación: ROJO confirmado; "
            "corrección sótanos y GPS. Decisión D1 demoler, prioridad inmediata."
        ),
    }


def plantilla_vacia_desde_precarga(precarga: dict[str, Any]) -> dict[str, Any]:
    """Precarga + validación/visita en Pendiente para nuevo vaciado."""
    out = dict(precarga)
    for k in [
        "val_edificio",
        "val_etiqueta",
        "val_geometria",
        "val_ranking",
        "estado_2da",
        "decision_D",
        "magnitud_M",
        "prioridad",
        "sistema",
        "ocupacion",
        "peligro_aledanos",
        "pct_columnas",
        "inclinacion",
        "dano_vigas",
        "riesgo_fachada",
    ]:
        out.setdefault(k, "Pendiente")
    out.setdefault("nombre_conf", precarga.get("nombre_hab", ""))
    out.setdefault("pisos_conf", str(precarga.get("pisos_f1", "")).split("/")[0].strip())
    return out


def enrich_workbook(
    xlsx_path,
    control_df: pd.DataFrame,
    indice_extra_rows: list[dict] | None = None,
):
    """Agrega solo Control_2da_ronda (+ listas) al Excel operativo de listados."""
    wb = load_workbook(xlsx_path)
    _add_list_sheet(wb)
    _write_df_sheet(wb, "Control_2da_ronda", control_df)

    # dropdowns on control sheet for key columns
    ws = wb["Control_2da_ronda"]
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    n = ws.max_row
    mapping = {
        "estado_2da_ronda": "estado_2da",
        "validacion_precarga": "si_no_parcial",
        "etiqueta_roja_confirmada": "si_no_insuf",
        "decision_D": "decision_D",
        "magnitud_M": "magnitud_M",
        "prioridad_operativa": "prioridad",
    }
    list_cols = {k: i + 1 for i, k in enumerate(LISTAS.keys())}
    for col_name, lista_key in mapping.items():
        if col_name not in headers or n < 2:
            continue
        c = headers[col_name]
        letter = get_column_letter(c)
        lc = get_column_letter(list_cols[lista_key])
        nvals = len(LISTAS[lista_key])
        formula = f"Listas_desplegables!${lc}$2:${lc}${nvals + 1}"
        ws.add_data_validation(_dv(formula, f"{letter}2:{letter}{n}"))

    # quitar hojas de vaciado si existían de una generación previa
    for obsolete in (
        "Vaciado_informe_digital",
        "Vaciado_ejemplo_FrancoMar",
        "Resumen_ejecutivo_ejemplo",
    ):
        if obsolete in wb.sheetnames:
            del wb[obsolete]

    if "Indice_pestanas" in wb.sheetnames and indice_extra_rows:
        ws = wb["Indice_pestanas"]
        start = ws.max_row + 1
        for i, row in enumerate(indice_extra_rows):
            ws.cell(start + i, 1, row.get("pestana", ""))
            ws.cell(start + i, 2, row.get("contenido", ""))
            ws.cell(start + i, 3, row.get("como_usarla", ""))

    wb.save(xlsx_path)
    return xlsx_path
