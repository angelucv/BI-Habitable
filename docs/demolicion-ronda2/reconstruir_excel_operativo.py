# -*- coding: utf-8 -*-
"""Reconstruye Excel operativo consolidado desde Habitable + cruce existente."""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from consolidar_excel_operativo import (  # noqa: E402
    XLSX,
    XLSX_ALT,
    _es_la_guaira_series,
    build_indice,
    build_ranking_unico,
    rename_control_ejecutivo,
    write_df_executive,
)
from excel_control_vaciado import (  # noqa: E402
    LISTAS,
    _add_list_sheet,
    _dv,
    build_control_2da_ronda,
)
from generar_entregables_demolicion import build_ranking_rojo  # noqa: E402

CSV = Path(r"C:\Users\PC\Downloads\habitable_inspecciones_2026-08-20_12-02-48.csv")


def main() -> None:
    path = XLSX if XLSX.exists() else XLSX_ALT
    cruce_ej = pd.read_excel(path, sheet_name="Cruce_informes", header=2)
    cruce = cruce_ej.rename(
        columns={"ID_Habitable": "hab_id", "Archivo_informe": "archivo_pdf"}
    )

    hab = pd.read_csv(CSV, low_memory=False)
    if "certificado" in hab.columns:
        hab["certificado"] = (
            hab["certificado"].astype(str).str.replace(r'^="?|"?$', "", regex=True)
        )
    hab_rojo = hab[hab["etiqueta"].astype(str).str.upper().str.contains("ROJO", na=False)].copy()

    ids: set[int] = set()
    for v in cruce.get("hab_id", []):
        try:
            if pd.notna(v) and str(v).strip() not in ("", "nan"):
                ids.add(int(float(v)))
        except (TypeError, ValueError):
            pass

    ranking = build_ranking_rojo(hab_rojo, ids)
    ranking_unico = build_ranking_unico(ranking)
    ranking_lg = ranking.loc[_es_la_guaira_series(ranking)].copy()
    ranking_lg = ranking_lg.sort_values("score_gravedad", ascending=False).reset_index(drop=True)
    ranking_lg["rank_gravedad_LaGuaira"] = ranking_lg.index + 1
    control = build_control_2da_ronda(ranking_lg, cruce)
    control_ej = rename_control_ejecutivo(control)

    resumen = pd.read_excel(path, sheet_name="Resumen", header=2)
    drop_pat = r"Ranking_ROJO|La Guaira \(filtro|Priorizar \(Alta|Top200|Control_2da"
    resumen = resumen[
        ~resumen["metrica"].astype(str).str.contains(drop_pat, na=False, regex=True)
    ]
    extra = pd.DataFrame(
        [
            {"metrica": "ROJO total (Ranking_ROJO)", "valor": len(ranking_unico)},
            {
                "metrica": "ROJO La Guaira (filtro Es_La_Guaira=Sí)",
                "valor": int((ranking_unico["Es_La_Guaira"] == "Sí").sum()),
            },
            {
                "metrica": "Priorizar (Alta/Muy alta)",
                "valor": int((ranking_unico["Prioridad_visita"] == "Priorizar").sum()),
            },
            {
                "metrica": "Top200",
                "valor": int((ranking_unico["Es_Top200"] == "Sí").sum()),
            },
            {"metrica": "Casos en Control_2da_ronda", "valor": len(control_ej)},
        ]
    )
    resumen2 = pd.concat([resumen, extra], ignore_index=True)
    catalogo = pd.read_excel(path, sheet_name="Catalogo_campos", header=2)
    indice = build_indice()

    try:
        with open(path, "a"):
            pass
        target = path
    except PermissionError:
        target = XLSX_ALT
        print(f"Archivo abierto; se guarda en {target.name}")

    wb = Workbook()
    wb.remove(wb.active)
    write_df_executive(
        wb,
        "Indice",
        indice,
        "LIBRO OPERATIVO — 2.ª RONDA ROJO / DEMOLICIÓN",
        "Índice. El vaciado tipo informe está en el Excel ejemplo Franco Mar (acompaña el Word).",
    )
    write_df_executive(
        wb,
        "Resumen",
        resumen2,
        "RESUMEN EJECUTIVO",
        "Indicadores del lote PDF y del universo ROJO (corte Habitable 20/08/2026).",
    )
    write_df_executive(
        wb,
        "Cruce_informes",
        cruce_ej,
        "CRUCE — INFORMES DETALLADOS × HABITABLE",
        "Una fila por PDF del lote. Revisar Calidad_cruce_Habitable.",
    )
    write_df_executive(
        wb,
        "Ranking_ROJO",
        ranking_unico,
        "RANKING ROJO — UNIVERSO COMPLETO",
        "FILTRAR: Prioridad_visita=Priorizar | Es_La_Guaira=Sí | Es_Top200=Sí | "
        "En_lote_PDF=Sí | Banda=Alta o Muy alta.",
    )
    write_df_executive(
        wb,
        "Control_2da_ronda",
        control_ej,
        "CONTROL 2.ª RONDA — COLA DE TRABAJO",
        "Actualizar Estado, Validación, Decisión D y Magnitud M (listas). Franco Mar = ejemplo D1.",
    )
    write_df_executive(
        wb,
        "Catalogo_campos",
        catalogo,
        "CATÁLOGO DE CAMPOS — FORMATO 2.ª RONDA",
        "Referencia de niveles A0–F alineada al Word de formato.",
    )
    _add_list_sheet(wb)

    ws = wb["Control_2da_ronda"]
    headers = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    n = ws.max_row
    mapping = {
        "Estado_2da_ronda": "estado_2da",
        "Validacion_precarga": "si_no_parcial",
        "Etiqueta_roja_confirmada": "si_no_insuf",
        "Decision_D": "decision_D",
        "Magnitud_M": "magnitud_M",
        "Prioridad_operativa": "prioridad",
    }
    list_cols = {k: i + 1 for i, k in enumerate(LISTAS.keys())}
    for col_name, lista_key in mapping.items():
        if col_name not in headers or n < 4:
            continue
        c = headers[col_name]
        letter = get_column_letter(c)
        lc = get_column_letter(list_cols[lista_key])
        nvals = len(LISTAS[lista_key])
        formula = f"Listas_desplegables!${lc}$2:${lc}${nvals + 1}"
        ws.add_data_validation(_dv(formula, f"{letter}4:{letter}{n}"))

    wb.save(target)
    print(f"OK {target}")
    print(f"Hojas: {wb.sheetnames}")
    print(
        "Priorizar=",
        int((ranking_unico["Prioridad_visita"] == "Priorizar").sum()),
        "LG=",
        int((ranking_unico["Es_La_Guaira"] == "Sí").sum()),
        "Control=",
        len(control_ej),
    )


if __name__ == "__main__":
    main()
