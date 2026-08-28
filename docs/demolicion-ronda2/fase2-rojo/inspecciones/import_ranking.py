"""Importación masiva desde Excel Ranking ROJO (Habitable)."""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterator

from django.db import transaction

from inspecciones import choices as ch
from inspecciones.models import CasoRojo

HOJAS_VALIDAS = (
    "Ranking_ROJO_LaGuaira",
    "Ranking_ROJO_nacional",
    "Ranking_ROJO_Top200",
)


def _txt(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _decimal_coord(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        d = Decimal(str(value))
        if abs(d) > 180:
            return None
        return d
    except (InvalidOperation, ValueError):
        return None


def _fecha(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def _riesgos(row: dict[str, Any]) -> str:
    partes = []
    for key, label in (
        ("riesgo_externo", "Ext"),
        ("riesgo_severo", "Sev"),
        ("riesgo_moderado", "Mod"),
        ("riesgo_componentes", "Comp"),
    ):
        v = _txt(row.get(key))
        if v:
            partes.append(f"{label} {v}")
    return " / ".join(partes)


def row_to_defaults(row: dict[str, Any], *, hoja: str) -> dict[str, Any]:
    hab_id = _int(row.get("id"))
    if not hab_id:
        raise ValueError("Fila sin id Habitable")

    municipio = _txt(row.get("municipio"))
    parroquia = _txt(row.get("parroquia"))
    muni_parr = " / ".join(p for p in (municipio, parroquia) if p)

    rank_lg = _txt(row.get("rank_gravedad_LaGuaira"))
    rank_nac = _txt(row.get("rank_gravedad"))
    if hoja == "Ranking_ROJO_LaGuaira" and rank_lg:
        puestos = f"La Guaira #{rank_lg}" + (f" / Nac #{rank_nac}" if rank_nac else "")
    elif rank_nac:
        puestos = f"Nacional #{rank_nac}"
    else:
        puestos = ""

    lat = _decimal_coord(row.get("lat"))
    lng = _decimal_coord(row.get("lng"))
    gps_hab = f"{lat}, {lng}" if lat is not None and lng is not None else ""

    num_pisos = row.get("num_pisos")
    pisos_f1 = _txt(num_pisos) if num_pisos not in (None, "") else ""

    defaults: dict[str, Any] = {
        "hab_id": hab_id,
        "certificado": _txt(row.get("certificado")),
        "nombre_hab": _txt(row.get("nombre_edificacion")),
        "etiqueta_f1": _txt(row.get("etiqueta")) or "ROJO",
        "fecha_f1": _fecha(row.get("created_at")),
        "inspector_f1": _txt(row.get("inspector_nombre")),
        "direccion_hab": _txt(row.get("direccion")),
        "muni_parr": muni_parr,
        "pisos_f1": pisos_f1,
        "riesgos_f1": _riesgos(row),
        "colapso_f1": _txt(row.get("ext_colapso_estructura")),
        "piso_crit_f1": _txt(row.get("piso_critico")),
        "acciones_f1": _txt(row.get("acc_medidas")),
        "obs_f1": _txt(row.get("observaciones")),
        "gps_hab": gps_hab,
        "lat": lat,
        "lng": lng,
        "score": _int(row.get("score_gravedad")),
        "banda": _txt(row.get("banda_prioridad")),
        "puestos": puestos,
        "score_detalle": _txt(row.get("score_detalle")),
        "prob_rel": _txt(row.get("prob_relativa_demolicion")),
        "uso": _txt(row.get("uso")),
    }
    return defaults


def iter_excel_rows(path: Path, hoja: str) -> Iterator[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f"Hoja «{hoja}» no encontrada. Disponibles: {', '.join(wb.sheetnames)}")
    ws = wb[hoja]
    rows = ws.iter_rows(values_only=True)
    headers = [(_txt(h) or f"col_{i}") for i, h in enumerate(next(rows))]
    for row in rows:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        yield dict(zip(headers, row))
    wb.close()


def filtrar_fila(
    row: dict[str, Any],
    *,
    min_score: int | None,
    bandas: set[str] | None,
    solo_con_gps: bool,
) -> bool:
    if min_score is not None:
        score = _int(row.get("score_gravedad"))
        if score is None or score < min_score:
            return False
    if bandas:
        banda = _txt(row.get("banda_prioridad"))
        if banda not in bandas:
            return False
    if solo_con_gps:
        if _decimal_coord(row.get("lat")) is None or _decimal_coord(row.get("lng")) is None:
            return False
    return True


@transaction.atomic
def importar_desde_excel(
    path: Path,
    *,
    hoja: str = "Ranking_ROJO_LaGuaira",
    limit: int | None = None,
    min_score: int | None = None,
    bandas: set[str] | None = None,
    solo_con_gps: bool = False,
    dry_run: bool = False,
    preservar_visitados: bool = True,
) -> dict[str, int]:
    """Importa o actualiza casos ROJO desde ranking Excel."""
    stats = {"leidas": 0, "creadas": 0, "actualizadas": 0, "omitidas": 0, "errores": 0}

    for row in iter_excel_rows(path, hoja):
        stats["leidas"] += 1
        if limit and stats["creadas"] + stats["actualizadas"] >= limit:
            break
        if not filtrar_fila(row, min_score=min_score, bandas=bandas, solo_con_gps=solo_con_gps):
            stats["omitidas"] += 1
            continue
        try:
            defaults = row_to_defaults(row, hoja=hoja)
            hab_id = defaults.pop("hab_id")
        except ValueError:
            stats["errores"] += 1
            continue

        if dry_run:
            if CasoRojo.objects.filter(hab_id=hab_id).exists():
                stats["actualizadas"] += 1
            else:
                stats["creadas"] += 1
            continue

        existente = CasoRojo.objects.filter(hab_id=hab_id).first()
        if existente and preservar_visitados:
            if existente.estado_2da not in ("", ch.Estado2daRonda.PENDIENTE):
                campos_precarga = {
                    k: v
                    for k, v in defaults.items()
                    if k
                    in {
                        "certificado",
                        "nombre_hab",
                        "score",
                        "banda",
                        "puestos",
                        "score_detalle",
                        "prob_rel",
                        "lat",
                        "lng",
                        "gps_hab",
                    }
                }
                for k, v in campos_precarga.items():
                    setattr(existente, k, v)
                existente.save(update_fields=list(campos_precarga.keys()))
                stats["actualizadas"] += 1
                continue

        _, created = CasoRojo.objects.update_or_create(hab_id=hab_id, defaults=defaults)
        if created:
            stats["creadas"] += 1
        else:
            stats["actualizadas"] += 1

    return stats
