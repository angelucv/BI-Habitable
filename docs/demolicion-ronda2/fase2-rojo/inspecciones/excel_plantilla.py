"""Generación de plantillas Excel — informe (vertical) y lote (filas)."""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable

from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from inspecciones.models import CasoRojo
from inspecciones.section_labels import CASE_FIELDSETS

PLANTILLA_VERSION = "v1.1"
INSTITUCION = "Comisión Presidencial para la Evaluación de Habitabilidad de Infraestructuras"
PROGRAMA = "Seguimiento inspecciones ROJO — Fase II"

LISTAS: dict[str, list[str]] = {
    "si_no_parcial": ["Sí", "No", "Parcial / corregir", "Pendiente"],
    "si_no_insuf": ["Sí", "No", "Insuficiente evidencia", "Pendiente"],
    "estado_2da": [
        "Pendiente verificación",
        "En visita",
        "Borrador",
        "Revisado",
        "Aprobado",
        "Publicado",
    ],
    "decision_D": [
        "D1 — Complementos requeridos",
        "D2 — Reparar / reconstruir",
        "D3 — Demoler",
        "D4 — Escombros / ya colapsado",
        "Pendiente",
    ],
    "magnitud_M": [
        "N/A (no es D2)",
        "M1 — Reparación local (menor intervención)",
        "M2 — Reparación importante",
        "M3 — Reconstrucción parcial",
        "M4 — Reconstrucción / refuerzo mayor (mayor intervención)",
        "Pendiente",
    ],
    "complementos_ejemplo": [
        "GEO,REI",
        "ENS,MOD",
        "MON,INV",
        "ALE,OTR",
    ],
    "prioridad": ["Inmediata", "Alta", "Programable", "Pendiente"],
    "abc": ["A", "B", "C", "No observable", "Pendiente"],
    "ocupacion": [
        "Desalojado",
        "Ocupación parcial",
        "Habitado irregularmente",
        "Escombros",
        "Pendiente",
    ],
    "sistema": [
        "Pórticos concreto armado",
        "Muros de concreto",
        "Mixto",
        "Acero",
        "Mampostería",
        "Otro",
        "Pendiente",
    ],
    "pct_columnas": ["No observable", "<10%", "10–30%", ">30%", ">50%", "Pendiente"],
    "inclinacion": [
        "No",
        "Sí — cualitativa",
        "Sí — con medición Δ",
        "No observable",
        "Pendiente",
    ],
    "peligro_aledanos": ["Sí", "No", "No observable", "Pendiente"],
}

LISTA_BY_FIELD: dict[str, str] = {
    "val_edificio": "si_no_parcial",
    "val_etiqueta": "si_no_insuf",
    "val_geometria": "si_no_parcial",
    "val_ranking": "si_no_parcial",
    "sistema": "sistema",
    "ocupacion": "ocupacion",
    "peligro_aledanos": "peligro_aledanos",
    "pct_columnas": "pct_columnas",
    "inclinacion": "inclinacion",
    "dano_vigas": "abc",
    "riesgo_fachada": "abc",
    "col_nivel": "abc",
    "vig_nivel": "abc",
    "mur_nivel": "abc",
    "mam_nivel": "abc",
    "repar_viable": "si_no_insuf",
    "estado_2da": "estado_2da",
    "decision_D": "decision_D",
    "magnitud_M": "magnitud_M",
    "prioridad": "prioridad",
}

PRECARGA_FIELDS = frozenset(
    {
        "hab_id",
        "certificado",
        "nombre_hab",
        "etiqueta_f1",
        "fecha_f1",
        "inspector_f1",
        "direccion_hab",
        "muni_parr",
        "pisos_f1",
        "riesgos_f1",
        "colapso_f1",
        "piso_crit_f1",
        "acciones_f1",
        "obs_f1",
        "gps_hab",
        "score",
        "banda",
        "puestos",
        "score_detalle",
        "prob_rel",
    }
)

VISITA_PENDIENTE: dict[str, str] = {
    "val_edificio": "Pendiente",
    "val_etiqueta": "Pendiente",
    "val_geometria": "Pendiente",
    "val_ranking": "Pendiente",
    "sistema": "Pendiente",
    "ocupacion": "Pendiente",
    "peligro_aledanos": "Pendiente",
    "pct_columnas": "Pendiente",
    "inclinacion": "Pendiente",
    "dano_vigas": "Pendiente",
    "riesgo_fachada": "Pendiente",
    "col_nivel": "Pendiente",
    "vig_nivel": "Pendiente",
    "mur_nivel": "Pendiente",
    "mam_nivel": "Pendiente",
    "repar_viable": "Pendiente",
    "estado_2da": "En visita",
    "decision_D": "Pendiente",
    "magnitud_M": "Pendiente",
    "prioridad": "Pendiente",
}

SKIP_EXPORT_FIELDS = frozenset(
    {"inspector_asignado", "revisor_asignado", "created_at", "updated_at", "lat", "lng"}
)

TALL_TEXT_FIELDS = frozenset(
    {
        "obs_f1",
        "score_detalle",
        "correcciones",
        "analisis_libre",
        "col_evidencia",
        "vig_evidencia",
        "escaleras",
        "preexistentes",
        "mam_diag",
        "proc_codigos",
        "proc_notas",
        "medidas",
        "justificacion",
        "resumen_ejecutivo",
    }
)


def _build_rows_def() -> list[tuple[str, str | None, str | None, str | None, bool]]:
    rows: list[tuple[str, str | None, str | None, str | None, bool]] = []
    for section_title, opts in CASE_FIELDSETS:
        if "Auditoría" in section_title:
            continue
        rows.append((section_title, None, None, None, False))
        for fname in opts["fields"]:
            if fname in SKIP_EXPORT_FIELDS:
                continue
            field = CasoRojo._meta.get_field(fname)
            label = str(field.verbose_name)
            lista = LISTA_BY_FIELD.get(fname)
            precarga = fname in PRECARGA_FIELDS
            rows.append((section_title, label, fname, lista, precarga))
    return rows


ROWS_DEF = _build_rows_def()
EXPORT_COLUMNS = [r[2] for r in ROWS_DEF if r[2]]


def _fmt_value(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, models.Model):
        return str(val)
    return val


def caso_to_dict(caso: CasoRojo, *, plantilla_visita: bool = True) -> dict[str, Any]:
    """Convierte caso a dict para Excel. Si plantilla_visita, limpia campos de visita no llenos."""
    data: dict[str, Any] = {}
    for col in EXPORT_COLUMNS:
        data[col] = _fmt_value(getattr(caso, col, ""))

    if plantilla_visita:
        for key, default in VISITA_PENDIENTE.items():
            cur = data.get(key, "")
            if cur in ("", "Pendiente", None):
                data[key] = default
        for key in EXPORT_COLUMNS:
            if key in PRECARGA_FIELDS or key in VISITA_PENDIENTE:
                continue
            if data.get(key) in ("", None):
                data[key] = ""

    data["_hab_id"] = caso.hab_id
    data["_nombre"] = caso.nombre_conf or caso.nombre_hab or ""
    return data


def _sheet_name_safe(hab_id: int, nombre: str = "") -> str:
    base = f"{hab_id}"
    if nombre:
        slug = re.sub(r"[^\w\s-]", "", nombre)[:12].strip().replace(" ", "_")
        if slug:
            base = f"{hab_id}_{slug}"
    return base[:31]


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _section_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _add_tricolor_bar(ws, row: int = 1) -> None:
    colors = ("FFCC00", "00247D", "CF142B")
    widths = (12, 12, 12)
    col = 1
    for color, w in zip(colors, widths):
        c = ws.cell(row, col, "")
        c.fill = PatternFill("solid", fgColor=color)
        ws.column_dimensions[get_column_letter(col)].width = w
        col += 1
    ws.row_dimensions[row].height = 6


def _write_portada(wb: Workbook, titulo: str, subtitulo: str, notas: list[tuple[str, str]]) -> None:
    ws = wb.active
    ws.title = "Portada"
    _add_tricolor_bar(ws, 1)
    ws["A3"] = titulo
    ws["A3"].font = Font(name="Calibri", size=16, bold=True, color="00247D")
    ws.merge_cells("A3:D3")
    ws["A4"] = INSTITUCION
    ws["A4"].font = Font(name="Calibri", size=10, color="475569")
    ws.merge_cells("A4:D4")
    ws["A5"] = subtitulo
    ws["A5"].font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
    ws.merge_cells("A5:D5")
    r = 7
    for tit, txt in notas:
        ws.cell(r, 1, tit).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="EEF2F8")
        ws.cell(r, 2, txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 48
        r += 1
    ws.cell(r, 1, "plantilla_version")
    ws.cell(r, 2, PLANTILLA_VERSION)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70


def _add_list_sheet(wb: Workbook) -> None:
    name = "Listas_desplegables"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for col_i, (key, values) in enumerate(LISTAS.items(), start=1):
        ws.cell(1, col_i, key)
        for row_i, val in enumerate(values, start=2):
            ws.cell(row_i, col_i, val)
    ws.sheet_state = "hidden"


def _dv(formula: str, sqref: str) -> DataValidation:
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valor no válido",
        error="Elija un valor de la lista.",
    )
    dv.add(sqref)
    return dv


def _write_informe_sheet(
    wb: Workbook,
    sheet_name: str,
    data: dict[str, Any],
    *,
    titulo: str | None = None,
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    thin = _thin_border()
    font_title = Font(name="Calibri", size=14, bold=True, color="00247D")
    font_sec = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_lab = Font(name="Calibri", size=10, bold=True)
    font_val = Font(name="Calibri", size=10)
    font_note = Font(name="Calibri", size=9, italic=True, color="666666")

    ws["A1"] = titulo or f"INFORME — ID {data.get('hab_id', data.get('_hab_id', ''))}"
    ws["A1"].font = font_title
    ws.merge_cells("A1:B1")
    ws["A2"] = (
        "Azul claro = precarga (validar). Amarillo = listas desplegables. "
        "Blanco = texto libre. Devuelva este archivo al coordinador para carga al sistema."
    )
    ws["A2"].font = font_note
    ws.merge_cells("A2:B2")

    dropdown_cells: list[tuple[int, str]] = []
    r = 4
    for seccion, etiqueta, key, lista_key, es_precarga in ROWS_DEF:
        if etiqueta is None:
            ws.cell(r, 1, seccion)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell = ws.cell(r, 1)
            cell.fill = _section_fill("2E75B6")
            cell.font = font_sec
            r += 1
            continue
        ws.cell(r, 1, etiqueta).font = font_lab
        ws.cell(r, 1).border = thin
        val = data.get(key, "")
        c = ws.cell(r, 2, val)
        c.font = font_val
        c.border = thin
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if es_precarga:
            c.fill = PatternFill("solid", fgColor="D9E2F3")
        if lista_key:
            dropdown_cells.append((r, lista_key))
            if not es_precarga:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
        if key in TALL_TEXT_FIELDS:
            ws.row_dimensions[r].height = 55
        r += 1

    list_cols = {k: i + 1 for i, k in enumerate(LISTAS.keys())}
    for row_i, lista_key in dropdown_cells:
        col_i = list_cols[lista_key]
        col_letter = get_column_letter(col_i)
        nvals = len(LISTAS[lista_key])
        formula = f"Listas_desplegables!${col_letter}$2:${col_letter}${nvals + 1}"
        ws.add_data_validation(_dv(formula, f"$B${row_i}"))

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 76
    ws.row_dimensions[1].height = 22


def _write_lote_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    name = "Lote_informes"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
    precarga_fill = PatternFill("solid", fgColor="D9E2F3")
    lista_fill = PatternFill("solid", fgColor="FFF2CC")

    labels = []
    for _sec, etiqueta, key, lista_key, es_precarga in ROWS_DEF:
        if key:
            labels.append((key, etiqueta, lista_key, es_precarga))

    for col_i, (key, label, _lk, _pre) in enumerate(labels, start=1):
        ws.cell(1, col_i, key).font = header_font
        ws.cell(1, col_i).fill = header_fill
        ws.cell(2, col_i, label).font = Font(bold=True, size=9)
        ws.column_dimensions[get_column_letter(col_i)].width = min(24, max(12, len(label) + 2))

    list_cols = {k: i + 1 for i, k in enumerate(LISTAS.keys())}
    for row_idx, data in enumerate(rows, start=3):
        for col_i, (key, _label, lista_key, es_precarga) in enumerate(labels, start=1):
            val = data.get(key, "")
            c = ws.cell(row_idx, col_i, val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if es_precarga:
                c.fill = precarga_fill
            elif lista_key:
                c.fill = lista_fill
                lk = list_cols.get(lista_key)
                if lk:
                    col_letter = get_column_letter(lk)
                    nvals = len(LISTAS[lista_key])
                    formula = f"Listas_desplegables!${col_letter}$2:${col_letter}${nvals + 1}"
                    ws.add_data_validation(_dv(formula, f"{get_column_letter(col_i)}{row_idx}"))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(labels))}{max(3, len(rows) + 2)}"


def generar_excel_informe(caso: CasoRojo) -> bytes:
    """Un edificio — hoja vertical tipo informe ejecutivo."""
    wb = Workbook()
    data = caso_to_dict(caso, plantilla_visita=True)
    nombre = data.get("_nombre") or data.get("nombre_hab") or ""
    _write_portada(
        wb,
        f"Plantilla informe — ID {caso.hab_id}",
        PROGRAMA,
        [
            ("Edificio", f"{nombre} (ID Habitable {caso.hab_id})"),
            ("Uso", "Complete la visita detallada (secciones 3–11). Precarga en azul — validar."),
            ("Devolución", "Remita este archivo al coordinador para validación, carga y PDF."),
            ("Formato", "Informe individual (1 hoja vertical por edificio)."),
        ],
    )
    _add_list_sheet(wb)
    sheet = _sheet_name_safe(caso.hab_id, nombre)
    _write_informe_sheet(wb, sheet, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_excel_informes_multiples(casos: Iterable[CasoRojo]) -> bytes:
    """Varios edificios — una hoja vertical por caso en el mismo libro."""
    casos_list = list(casos)
    wb = Workbook()
    ids = ", ".join(str(c.hab_id) for c in casos_list[:8])
    if len(casos_list) > 8:
        ids += f" … (+{len(casos_list) - 8})"
    _write_portada(
        wb,
        f"Plantillas informe — {len(casos_list)} caso(s)",
        PROGRAMA,
        [
            ("Casos incluidos", ids),
            ("Uso", "Una hoja por edificio. Complete visita en cada pestaña."),
            ("Devolución", "Remita el libro completo al coordinador."),
            ("Formato", "Informe múltiple (pestaña por edificio)."),
        ],
    )
    _add_list_sheet(wb)
    for caso in casos_list:
        data = caso_to_dict(caso, plantilla_visita=True)
        nombre = data.get("_nombre") or ""
        sheet = _sheet_name_safe(caso.hab_id, nombre)
        _write_informe_sheet(wb, sheet, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_excel_lote(casos: Iterable[CasoRojo]) -> bytes:
    """Varios edificios — una fila por caso."""
    casos_list = list(casos)
    rows = [caso_to_dict(c, plantilla_visita=True) for c in casos_list]
    wb = Workbook()
    _write_portada(
        wb,
        f"Plantilla lote — {len(casos_list)} caso(s)",
        PROGRAMA,
        [
            ("Formato", "Una fila por edificio en hoja «Lote_informes»."),
            ("Columnas", "Fila 1 = clave técnica · Fila 2 = etiqueta · Desde fila 3 = datos."),
            ("Precarga", "Columnas azules en informe; aquí también vienen precargadas."),
            ("Devolución", "Remita al coordinador para validación masiva y PDF por caso."),
        ],
    )
    _add_list_sheet(wb)
    _write_lote_sheet(wb, rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nombre_archivo_informe(caso: CasoRojo) -> str:
    slug = re.sub(r"[^\w.-]+", "_", (caso.nombre_conf or caso.nombre_hab or "caso")[:30])
    return f"informe-rojo-{caso.hab_id}-{slug}.xlsx"


def nombre_archivo_lote(n: int) -> str:
    return f"lote-informes-rojo-{n}-casos.xlsx"


def nombre_archivo_multinforme(n: int) -> str:
    return f"informes-rojo-{n}-pestanas.xlsx"
