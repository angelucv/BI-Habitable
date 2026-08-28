# -*- coding: utf-8 -*-
"""Consolida rankings redundantes y aplica formato ejecutivo al Excel operativo."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_control_vaciado import (
    LISTAS,
    _add_list_sheet,
    _dv,
    build_control_2da_ronda,
    enrich_workbook,
)

OUT = Path(__file__).resolve().parent
XLSX = OUT / "cruce-informes-demolicion-habitable-2026-08-20.xlsx"
XLSX_ALT = OUT / "cruce-informes-demolicion-habitable-2026-08-20-v2.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=9)
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
ZEBRA = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _es_la_guaira(row: pd.Series) -> bool:
    est = str(row.get("estado", "") or "").upper()
    mun = str(row.get("municipio", "") or "").upper()
    # también aceptar columna Estado/Municipio ya renombrada
    if not est or est == "NAN":
        est = str(row.get("Estado", "") or "").upper()
    if not mun or mun == "NAN":
        mun = str(row.get("Municipio", "") or "").upper()
    return ("GUAIRA" in est) or ("VARGAS" in est) or ("GUAIRA" in mun) or ("VARGAS" in mun)


def _es_la_guaira_series(df: pd.DataFrame) -> pd.Series:
    est = df.get("estado", df.get("Estado", pd.Series("", index=df.index))).astype(str).str.upper()
    mun = df.get("municipio", df.get("Municipio", pd.Series("", index=df.index))).astype(str).str.upper()
    return (
        est.str.contains("GUAIRA", na=False)
        | est.str.contains("VARGAS", na=False)
        | mun.str.contains("GUAIRA", na=False)
        | mun.str.contains("VARGAS", na=False)
    )


def normalize_ranking_raw(df: pd.DataFrame) -> pd.DataFrame:
    """Acepta hoja cruda o ya ejecutiva y devuelve columnas canónicas internas."""
    ren = {
        "Puesto_nacional": "rank_gravedad",
        "Score": "score_gravedad",
        "Banda": "banda_prioridad",
        "ID_Habitable": "id",
        "Edificacion": "nombre_edificacion",
        "Municipio": "municipio",
        "Parroquia": "parroquia",
        "Estado": "estado",
        "Direccion": "direccion",
        "Pisos": "num_pisos",
        "En_lote_PDF": "en_lote_informes_pdf",
        "Piso_critico_F1": "piso_critico",
        "Riesgo_externo": "riesgo_externo",
        "Riesgo_severo": "riesgo_severo",
        "Colapso_estructura": "ext_colapso_estructura",
        "Acciones_Fase1": "acc_medidas",
        "Detalle_score": "score_detalle",
        "Probabilidad_relativa": "prob_relativa_demolicion",
        "Certificado": "certificado",
        "Inspector_Fase1": "inspector_nombre",
        "Fecha_Fase1": "created_at",
        "Observaciones_Fase1": "observaciones",
        "Lat": "lat",
        "Lng": "lng",
    }
    out = df.rename(columns={k: v for k, v in ren.items() if k in df.columns and v not in df.columns})
    required = ["rank_gravedad", "score_gravedad", "id", "nombre_edificacion"]
    missing = [c for c in required if c not in out.columns]
    if missing:
        raise SystemExit(f"Ranking sin columnas requeridas: {missing}")
    return out


def build_ranking_unico(ranking: pd.DataFrame) -> pd.DataFrame:
    """Una sola hoja: todos los ROJO + flags para filtrar en Excel."""
    df = normalize_ranking_raw(ranking)
    df["Es_La_Guaira"] = _es_la_guaira_series(df).map({True: "Sí", False: "No"})

    lg = df[df["Es_La_Guaira"] == "Sí"].sort_values(
        ["score_gravedad", "id"], ascending=[False, True]
    )
    puesto_lg = {int(i): k + 1 for k, i in enumerate(lg["id"].tolist())}
    df["Puesto_LaGuaira"] = df["id"].map(
        lambda x: puesto_lg.get(int(x), "") if pd.notna(x) else ""
    )

    df["Es_Top200"] = (df["rank_gravedad"] <= 200).map({True: "Sí", False: "No"})
    df["Prioridad_visita"] = df["banda_prioridad"].map(
        {
            "Muy alta": "Priorizar",
            "Alta": "Priorizar",
            "Media": "Revisar",
            "Baja-media": "Cola general",
            "Baja / datos incompletos": "Curar dato / cola general",
        }
    ).fillna("Cola general")

    # nombres ejecutivos
    out = pd.DataFrame(
        {
            "Puesto_nacional": df["rank_gravedad"],
            "Score": df["score_gravedad"],
            "Banda": df["banda_prioridad"],
            "Prioridad_visita": df["Prioridad_visita"],
            "Es_La_Guaira": df["Es_La_Guaira"],
            "Puesto_LaGuaira": df["Puesto_LaGuaira"],
            "Es_Top200": df["Es_Top200"],
            "En_lote_PDF": df.get("en_lote_informes_pdf", ""),
            "Edificacion": df["nombre_edificacion"],
            "Municipio": df["municipio"],
            "Parroquia": df.get("parroquia", ""),
            "Estado": df["estado"],
            "Direccion": df["direccion"],
            "Pisos": df["num_pisos"],
            "Piso_critico_F1": df.get("piso_critico", ""),
            "Riesgo_externo": df.get("riesgo_externo", ""),
            "Riesgo_severo": df.get("riesgo_severo", ""),
            "Colapso_estructura": df.get("ext_colapso_estructura", ""),
            "Acciones_Fase1": df.get("acc_medidas", ""),
            "Probabilidad_relativa": df.get("prob_relativa_demolicion", ""),
            "Detalle_score": df.get("score_detalle", ""),
            "ID_Habitable": df["id"],
            "Certificado": df.get("certificado", ""),
            "Inspector_Fase1": df.get("inspector_nombre", ""),
            "Fecha_Fase1": df.get("created_at", ""),
            "Lat": df.get("lat", ""),
            "Lng": df.get("lng", ""),
            "Observaciones_Fase1": df.get("observaciones", ""),
        }
    )
    return out.sort_values(["Score", "Puesto_nacional"], ascending=[False, True]).reset_index(
        drop=True
    )


def rename_cruce_ejecutivo(df: pd.DataFrame) -> pd.DataFrame:
    mapa = {
        "archivo_pdf": "Archivo_informe",
        "tipo_informe": "Tipo_informe",
        "nombre_edificacion_informe": "Edificacion_informe",
        "ubicacion_informe": "Ubicacion_informe",
        "fecha_inspeccion_informe": "Fecha_informe",
        "recomienda_demolicion": "Senal_demolicion_en_PDF",
        "dictamen_etiqueta": "Dictamen_en_PDF",
        "match_calidad": "Calidad_cruce_Habitable",
        "hab_id": "ID_Habitable",
        "hab_certificado": "Certificado",
        "hab_etiqueta": "Etiqueta_Habitable",
        "hab_nombre": "Edificacion_Habitable",
        "hab_municipio": "Municipio",
        "hab_parroquia": "Parroquia",
        "hab_num_pisos": "Pisos_Habitable",
        "hab_direccion": "Direccion_Habitable",
        "evaluadores": "Evaluadores_PDF",
        "num_pisos_informe": "Pisos_informe",
    }
    cols = [c for c in [
        "Archivo_informe", "Tipo_informe", "Edificacion_informe", "Ubicacion_informe",
        "Fecha_informe", "Senal_demolicion_en_PDF", "Dictamen_en_PDF",
        "Calidad_cruce_Habitable", "ID_Habitable", "Certificado", "Etiqueta_Habitable",
        "Edificacion_Habitable", "Municipio", "Parroquia", "Pisos_Habitable",
        "Direccion_Habitable", "Evaluadores_PDF", "Pisos_informe", "match_score",
        "match_alternativos", "extracto_conclusiones",
    ] if c in df.rename(columns=mapa).columns or c in mapa.values()]
    ren = df.rename(columns=mapa)
    keep = [c for c in [
        "Archivo_informe", "Tipo_informe", "Edificacion_informe", "Ubicacion_informe",
        "Fecha_informe", "Senal_demolicion_en_PDF", "Dictamen_en_PDF",
        "Calidad_cruce_Habitable", "match_score", "ID_Habitable", "Certificado",
        "Etiqueta_Habitable", "Edificacion_Habitable", "Municipio", "Parroquia",
        "Pisos_Habitable", "Direccion_Habitable", "Evaluadores_PDF", "Pisos_informe",
        "match_alternativos", "extracto_conclusiones",
    ] if c in ren.columns]
    return ren[keep]


def rename_control_ejecutivo(df: pd.DataFrame) -> pd.DataFrame:
    mapa = {
        "rank_gravedad_LaGuaira": "Puesto_LaGuaira",
        "rank_gravedad": "Puesto_nacional",
        "score_gravedad": "Score",
        "banda_prioridad": "Banda",
        "prob_relativa_demolicion": "Probabilidad_relativa",
        "en_lote_informes_pdf": "En_lote_PDF",
        "estado_2da_ronda": "Estado_2da_ronda",
        "validacion_precarga": "Validacion_precarga",
        "etiqueta_roja_confirmada": "Etiqueta_roja_confirmada",
        "decision_D": "Decision_D",
        "magnitud_M": "Magnitud_M",
        "prioridad_operativa": "Prioridad_operativa",
        "fecha_visita_2": "Fecha_visita_2",
        "evaluador_visita_2": "Evaluador_visita_2",
        "supervisor_visita_2": "Supervisor_visita_2",
        "id": "ID_Habitable",
        "certificado": "Certificado",
        "nombre_edificacion": "Edificacion",
        "direccion": "Direccion",
        "municipio": "Municipio",
        "parroquia": "Parroquia",
        "num_pisos": "Pisos",
        "piso_critico": "Piso_critico_F1",
        "riesgo_externo": "Riesgo_externo",
        "riesgo_severo": "Riesgo_severo",
        "ext_colapso_estructura": "Colapso_estructura",
        "acc_medidas": "Acciones_Fase1",
        "score_detalle": "Detalle_score",
        "correcciones_precarga": "Correcciones_precarga",
        "resumen_ejecutivo": "Resumen_ejecutivo",
        "n_fotos": "N_fotos",
        "inspector_nombre": "Inspector_Fase1",
        "created_at": "Fecha_Fase1",
    }
    ren = df.rename(columns=mapa)
    order = [
        "Puesto_LaGuaira", "Puesto_nacional", "Score", "Banda", "En_lote_PDF",
        "Estado_2da_ronda", "Validacion_precarga", "Etiqueta_roja_confirmada",
        "Decision_D", "Magnitud_M", "Prioridad_operativa",
        "Edificacion", "Municipio", "Parroquia", "Pisos",
        "Fecha_visita_2", "Evaluador_visita_2", "Supervisor_visita_2",
        "Resumen_ejecutivo", "Correcciones_precarga",
        "ID_Habitable", "Certificado", "Direccion",
        "Piso_critico_F1", "Riesgo_externo", "Riesgo_severo", "Colapso_estructura",
        "Acciones_Fase1", "Detalle_score", "Probabilidad_relativa",
        "N_fotos", "Inspector_Fase1", "Fecha_Fase1", "lat", "lng",
    ]
    keep = [c for c in order if c in ren.columns]
    return ren[keep]


def style_sheet_executive(ws, title: str | None = None, note: str | None = None):
    """Formato ejecutivo: título opcional, encabezado, filtro, freeze, anchos."""
    # if title, insert rows at top — caller should already have title rows
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    # detect header row: first row with values, or row 2 if A1 is title spanning
    header_row = 1
    if ws["A1"].value and ws["B1"].value is None and title:
        header_row = 2
    if note and ws.cell(2, 1).value and "Filtrar" in str(ws.cell(2, 1).value):
        header_row = 3

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for r in range(header_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.font = CELL_FONT
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if (r - header_row) % 2 == 0:
                if not cell.fill or cell.fill.fgColor is None or cell.fill.fgColor.rgb in (
                    "00000000",
                    None,
                ):
                    cell.fill = ZEBRA

    ws.freeze_panes = f"A{header_row + 1}"
    last_col = get_column_letter(max_col)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{max_row}"
    ws.row_dimensions[header_row].height = 32

    for i in range(1, max_col + 1):
        letter = get_column_letter(i)
        header = ws.cell(header_row, i).value
        width = 14
        if header:
            h = str(header)
            if "Edificacion" in h or "Direccion" in h or "Resumen" in h or "Detalle" in h:
                width = 32
            elif "Probabilidad" in h or "Observaciones" in h or "extracto" in h.lower():
                width = 36
            elif len(h) > 18:
                width = 20
        ws.column_dimensions[letter].width = width


def write_df_executive(
    wb,
    name: str,
    df: pd.DataFrame,
    titulo: str,
    nota_filtro: str,
):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = titulo
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, max(2, len(df.columns))))
    ws["A2"] = nota_filtro
    ws["A2"].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(8, max(2, len(df.columns))))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)

    style_sheet_executive(ws, title=titulo, note=nota_filtro)
    # fix header detection: we know header is row 3
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(df.shape[1])}{ws.max_row}"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 34
    return ws


def build_indice() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Hoja": "Indice",
                "Contenido": "Guía del libro operativo.",
                "Uso": "Empezar aquí.",
            },
            {
                "Hoja": "Resumen",
                "Contenido": "Indicadores del lote PDF y del universo ROJO.",
                "Uso": "Vista ejecutiva rápida.",
            },
            {
                "Hoja": "Cruce_informes",
                "Contenido": "Informes detallados del lote PDF cruzados con Habitable.",
                "Uso": "Seguimiento del lote ya visitado en detalle.",
            },
            {
                "Hoja": "Ranking_ROJO",
                "Contenido": "Todos los ROJO con score, banda y flags de filtro "
                "(La Guaira, Top200, Prioridad_visita, En_lote_PDF).",
                "Uso": "Usar Autofiltro: Prioridad_visita=Priorizar; Es_La_Guaira=Sí; Es_Top200=Sí.",
            },
            {
                "Hoja": "Control_2da_ronda",
                "Contenido": "Cola de trabajo 2.ª visita: validación, decisión D1–D5, magnitud M.",
                "Uso": "Actualizar estado y dictamen; listas desplegables en columnas de control.",
            },
            {
                "Hoja": "Catalogo_campos",
                "Contenido": "Campos del formato de inspección detallada (A0–F).",
                "Uso": "Referencia de diseño / sistema.",
            },
            {
                "Hoja": "(archivo aparte) Ejemplo vaciado… Franco Mar.xlsx",
                "Contenido": "Vaciado digital tipo informe; acompaña el Word de formato.",
                "Uso": "Capacitación / propuesta metodológica.",
            },
        ]
    )


def main():
    path = XLSX if XLSX.exists() else XLSX_ALT
    try:
        with open(path, "a"):
            pass
        target = path
    except PermissionError:
        target = XLSX_ALT
        print(f"Archivo abierto; se guarda en {target.name}")
        if not path.exists():
            raise
        # copy base if alt missing
        if not target.exists():
            import shutil
            shutil.copy2(path, target)

    # leer piezas existentes
    xl = pd.ExcelFile(path if path.exists() else target)
    sheets = xl.sheet_names

    if "Ranking_ROJO_nacional" in sheets:
        ranking_raw = pd.read_excel(path, sheet_name="Ranking_ROJO_nacional")
    elif "Ranking_ROJO" in sheets:
        # already consolidated — rebuild flags from executive cols if needed
        ranking_raw = pd.read_excel(path, sheet_name="Ranking_ROJO")
        # map back minimal for control if needed
    else:
        raise SystemExit("No se encontró hoja de ranking en el Excel.")

    # normalizar siempre
    ranking_canon = normalize_ranking_raw(ranking_raw)
    ranking_unico = build_ranking_unico(ranking_canon)
    ranking_lg_src = ranking_canon

    cruce = pd.read_excel(path, sheet_name="Cruce_informes")
    cruce_ej = rename_cruce_ejecutivo(cruce)

    ranking_lg = ranking_lg_src.loc[_es_la_guaira_series(ranking_lg_src)].copy()
    ranking_lg = ranking_lg.sort_values("score_gravedad", ascending=False).reset_index(drop=True)
    ranking_lg["rank_gravedad_LaGuaira"] = ranking_lg.index + 1
    print(f"ROJO La Guaira para control: {len(ranking_lg)}")
    control = build_control_2da_ronda(ranking_lg, cruce)
    control_ej = rename_control_ejecutivo(control)

    # resumen
    if "Resumen" in sheets:
        resumen = pd.read_excel(path, sheet_name="Resumen")
    else:
        resumen = pd.DataFrame([{"metrica": "ROJO total", "valor": len(ranking_unico)}])

    # enriquecer resumen
    extra = pd.DataFrame(
        [
            {"metrica": "ROJO total (Ranking_ROJO)", "valor": len(ranking_unico)},
            {
                "metrica": "ROJO La Guaira (filtro Es_La_Guaira=Sí)",
                "valor": int((ranking_unico["Es_La_Guaira"] == "Sí").sum()),
            },
            {
                "metrica": "Priorizar (banda Alta/Muy alta)",
                "valor": int((ranking_unico["Prioridad_visita"] == "Priorizar").sum()),
            },
            {
                "metrica": "Top200 nacional",
                "valor": int((ranking_unico["Es_Top200"] == "Sí").sum()),
            },
            {"metrica": "Casos en Control_2da_ronda", "valor": len(control_ej)},
        ]
    )
    resumen2 = pd.concat([resumen, extra], ignore_index=True)

    if "Catalogo_campos_propuestos" in sheets:
        catalogo = pd.read_excel(path, sheet_name="Catalogo_campos_propuestos")
    elif "Catalogo_campos" in sheets:
        catalogo = pd.read_excel(path, sheet_name="Catalogo_campos")
    else:
        catalogo = pd.DataFrame()

    indice = build_indice()

    # escribir libro nuevo limpio
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    write_df_executive(
        wb,
        "Indice",
        indice,
        "LIBRO OPERATIVO — 2.ª RONDA ROJO / DEMOLICIÓN",
        "Índice de hojas. El vaciado tipo informe está en el Excel ejemplo (Franco Mar), junto al Word.",
    )
    write_df_executive(
        wb,
        "Resumen",
        resumen2,
        "RESUMEN EJECUTIVO",
        "Indicadores del lote de informes y del universo ROJO (corte Habitable 20/08/2026).",
    )
    write_df_executive(
        wb,
        "Cruce_informes",
        cruce_ej,
        "CRUCE — INFORMES DETALLADOS × HABITABLE",
        "Una fila por PDF del lote. Revisar Calidad_cruce_Habitable.",
    )
    write_df_executive(
        wb,
        "Ranking_ROJO",
        ranking_unico,
        "RANKING ROJO — UNIVERSO COMPLETO",
        "FILTRAR con Autofiltro: Prioridad_visita = Priorizar | Es_La_Guaira = Sí | "
        "Es_Top200 = Sí | En_lote_PDF = Sí | Banda = Alta o Muy alta.",
    )
    write_df_executive(
        wb,
        "Control_2da_ronda",
        control_ej,
        "CONTROL 2.ª RONDA — COLA DE TRABAJO",
        "Actualizar Estado, Validación, Decisión D y Magnitud M (listas desplegables). "
        "Franco Mar figura como ejemplo D1.",
    )
    if len(catalogo):
        write_df_executive(
            wb,
            "Catalogo_campos",
            catalogo,
            "CATÁLOGO DE CAMPOS — FORMATO 2.ª RONDA",
            "Referencia de niveles A0–F alineada al Word de formato.",
        )

    _add_list_sheet(wb)

    # dropdowns control
    ws = wb["Control_2da_ronda"]
    # header is row 3
    headers = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    n = ws.max_row
    mapping = {
        "Estado_2da_ronda": "estado_2da",
        "Validacion_precarga": "si_no_parcial",
        "Etiqueta_roja_confirmada": "si_no_insuf",
        "Decision_D": "decision_D",
        "Magnitud_M": "magnitud_M",
        "Prioridad_operativa": "prioridad",
    }
    list_cols = {k: i + 1 for i, k in enumerate(LISTAS.keys())}
    for col_name, lista_key in mapping.items():
        if col_name not in headers or n < 4:
            continue
        c = headers[col_name]
        letter = get_column_letter(c)
        lc = get_column_letter(list_cols[lista_key])
        nvals = len(LISTAS[lista_key])
        formula = f"Listas_desplegables!${lc}$2:${lc}${nvals + 1}"
        ws.add_data_validation(_dv(formula, f"{letter}4:{letter}{n}"))

    wb.save(target)
    print(f"Excel operativo consolidado: {target}")
    print(f"Ranking_ROJO filas={len(ranking_unico)} | Control={len(control_ej)}")
    print("Hojas:", wb.sheetnames)


if __name__ == "__main__":
    main()
