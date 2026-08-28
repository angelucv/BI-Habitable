# -*- coding: utf-8 -*-
"""Genera Excel didáctico de vaciado (acompaña el Word de formato)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_control_vaciado import (
    LISTAS,
    _add_list_sheet,
    _dv,
    francomar_precarga,
    francomar_visita,
    plantilla_vacia_desde_precarga,
    write_vaciado_informe_sheet,
)

OUT = Path(__file__).resolve().parent
OUT_XLSX = OUT / "Ejemplo vaciado digital inspeccion 2da ronda Franco Mar.xlsx"


def _portada(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Portada_uso"
    ws["A1"] = (
        "EJEMPLO DE VACIADO DIGITAL — INSPECCIÓN DETALLADA (2.ª RONDA)"
    )
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws.merge_cells("A1:B1")

    filas = [
        (
            "Para qué es este archivo",
            "Demostrar cómo se llena el formato propuesto de segunda visita, "
            "con precarga de la inspección Habitable (Fase 1), ranking de prioridad "
            "y dictamen D1–D5 / M1–M4. Acompaña el documento Word explicativo del formato.",
        ),
        (
            "Documento Word asociado",
            "Formato propuesto inspeccion detallada verificacion ROJO.docx",
        ),
        (
            "Qué NO es este archivo",
            "No es el listado operativo de todos los ROJOS ni el cruce de informes PDF. "
            "Eso vive en el Excel de trabajo «cruce-informes-demolicion-habitable-…».",
        ),
        (
            "Hojas de este libro",
            "1) Portada_uso (esta) · 2) Vaciado_ejemplo_FrancoMar (lleno) · "
            "3) Vaciado_plantilla_en_blanco (para practicar) · "
            "4) Resumen_ejecutivo_ejemplo · 5) Listas_desplegables (oculta).",
        ),
        (
            "Caso de ejemplo",
            "Edificio Franco Mar / Edif. Francomar (Caraballeda, La Guaira). "
            "ID Habitable 171818. Fase 1 ROJO 20/07/2026. Ranking score 51 (Media). "
            "Visita detallada 11/08/2026 → decisión D1 Demoler.",
        ),
        (
            "Cómo leer los colores",
            "Azul claro = dato precargado de Fase 1 / ranking (validar). "
            "Amarillo = campo de la visita 2 con lista desplegable. "
            "Blanco = texto libre o dato abierto.",
        ),
        (
            "Uso sugerido",
            "Abrir el Word para la lógica del formato; abrir esta hoja de ejemplo "
            "para ver el vaciado concreto; copiar la plantilla en blanco si se quiere "
            "ensayar otro caso.",
        ),
    ]
    r = 3
    for titulo, texto in filas:
        ws.cell(r, 1, titulo).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="D6DCE4")
        ws.cell(r, 2, texto).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 55
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 92


def _resumen_ejecutivo(wb: Workbook) -> None:
    ws = wb.create_sheet("Resumen_ejecutivo_ejemplo")
    ws["A1"] = "RESUMEN EJECUTIVO — EJEMPLO FRANCO MAR"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws.merge_cells("A1:B1")
    filas = [
        ("Edificación", "Franco Mar / Edif. Francomar (Caraballeda, La Guaira)"),
        ("ID Habitable / certificado", "171818 / 73420200720261338"),
        ("Fase 1", "ROJO · 20/07/2026 · Inspector M. Nieves Duarte"),
        ("Ranking", "Score 51 · Banda Media · ~puesto 115 La Guaira / 264 nacional"),
        ("Visita detallada", "11/08/2026 · Burgos / García · Supervisión A. Quintero"),
        ("Validación precarga", "ROJO confirmado · Corrección sótanos 1→2 y GPS/dirección"),
        ("Hallazgo clave", "Piso crítico Nivel 1 · >50% columnas graves · pérdida de verticalidad"),
        ("Decisión", "D1 — Demoler · Prioridad inmediata · Magnitud M N/A"),
        ("Texto ejecutivo", francomar_visita()["resumen_ejecutivo"]),
        (
            "Próxima acción",
            "Proyecto de demolición controlada; mantener exclusión y monitoreo de vecinos",
        ),
    ]
    r = 3
    for a, b in filas:
        ws.cell(r, 1, a).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(r, 2, b).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 40 if a.startswith("Texto") else 20
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90


def main() -> None:
    wb = Workbook()
    _portada(wb)
    _add_list_sheet(wb)

    write_vaciado_informe_sheet(
        wb,
        "Vaciado_ejemplo_FrancoMar",
        francomar_precarga(),
        francomar_visita(),
        titulo=(
            "VACIADO DIGITAL — EJEMPLO LLENO (Franco Mar) · "
            "Precarga Fase 1 + ranking + visita 11/08/2026"
        ),
    )

    # plantilla en blanco (sin precarga de un caso real)
    write_vaciado_informe_sheet(
        wb,
        "Vaciado_plantilla_en_blanco",
        plantilla_vacia_desde_precarga(
            {
                "hab_id": "",
                "certificado": "",
                "nombre_hab": "",
                "etiqueta_f1": "ROJO",
                "fecha_f1": "",
                "inspector_f1": "",
                "direccion_hab": "",
                "muni_parr": "",
                "pisos_f1": "",
                "riesgos_f1": "",
                "colapso_f1": "",
                "piso_crit_f1": "",
                "acciones_f1": "",
                "obs_f1": "",
                "gps_hab": "",
                "score": "",
                "banda": "",
                "puestos": "",
                "score_detalle": "",
                "prob_rel": "",
            }
        ),
        None,
        titulo=(
            "VACIADO DIGITAL — PLANTILLA EN BLANCO · "
            "Pegue precarga Habitable/ranking y complete la visita 2"
        ),
    )

    _resumen_ejecutivo(wb)

    # orden de hojas
    order = [
        "Portada_uso",
        "Vaciado_ejemplo_FrancoMar",
        "Vaciado_plantilla_en_blanco",
        "Resumen_ejecutivo_ejemplo",
        "Listas_desplegables",
    ]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(OUT_XLSX)
    print(f"Excel ejemplo generado: {OUT_XLSX}")


if __name__ == "__main__":
    main()
